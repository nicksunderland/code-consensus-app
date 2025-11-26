import os
import csv
import re
from typing import List, Dict, Any, Optional
import xml.etree.ElementTree as ET
import sys
import openpyxl

sys.setrecursionlimit(5000)


# ==========================================================
# 0. GLOBAL CONFIG & UTILS
# ==========================================================

def normalize_code(code: str) -> str:
    if not code: return ""
    clean = re.sub(r'^(BLOCK|CHAPTER)\s+', '', code, flags=re.IGNORECASE)
    return clean.replace(".", "").strip().upper()


def strip_code_from_desc(raw_code: str, description: str) -> str:
    if not description or not raw_code: return description
    desc_clean = description
    desc_upper = description.upper()

    raw_prefix = f"{raw_code}".upper()
    if desc_upper.startswith(raw_prefix):
        desc_clean = desc_clean[len(raw_prefix):]

    clean_code = re.sub(r'^(BLOCK|CHAPTER)\s+', '', raw_code, flags=re.IGNORECASE).strip()
    clean_code_nodots = clean_code.replace(".", "")

    prefixes = [clean_code, clean_code_nodots, f"Chapter {clean_code}", f"Block {clean_code}"]
    if len(clean_code_nodots) > 3:
        dotted = f"{clean_code_nodots[:3]}.{clean_code_nodots[3:]}"
        prefixes.append(dotted)

    prefixes.sort(key=len, reverse=True)

    for p in prefixes:
        pattern = r"^" + re.escape(p) + r"[\s.:-]*"
        match = re.match(pattern, desc_clean, flags=re.IGNORECASE)
        if match:
            desc_clean = desc_clean[match.end():].strip()
            break

    return desc_clean.lstrip(" :.-").strip()


def get_common_description(descriptions: List[str]) -> str:
    if not descriptions: return "Category"
    if len(descriptions) == 1: return descriptions[0].strip()
    s_list = sorted(descriptions)
    s1, s2 = s_list[0], s_list[-1]
    match_len = 0
    for i in range(min(len(s1), len(s2))):
        if s1[i].upper() == s2[i].upper():
            match_len += 1
        else:
            break
    common = s1[:match_len]
    if match_len < len(s1) and s1[match_len].isalpha() and match_len > 0 and common[-1].isalpha():
        last_space = common.rfind(' ')
        if last_space > 0: common = common[:last_space]
    common = common.strip(" ,-:")
    if len(common) < 3: return "Category"
    return f"{common}"


# --- MAPPINGS ---
ROMAN_MAP = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5,
    "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10,
    "XI": 11, "XII": 12, "XIII": 13, "XIV": 14, "XV": 15,
    "XVI": 16, "XVII": 17, "XVIII": 18, "XIX": 19, "XX": 20,
    "XXI": 21, "XXII": 22
}


def custom_sort_key(code: str):
    if code in ROMAN_MAP:
        return (0, ROMAN_MAP[code], "")

    # Sort range codes (like "060-080") after Roman numerals but before regular codes
    if "-" in code:
        # Extract first number for sorting ranges
        match = re.match(r'^(\d+)', code)
        if match:
            return (1, int(match.group(1)), code)
        return (1, 0, code)

    # Sort regular numeric codes
    if code.isdigit():
        return (2, int(code), "")

    # Everything else (alphanumeric codes like ICD-10)
    return (3, 0, code)


ICD10_CHAPTER_LOOKUP = {
    "A": "I", "B": "I", "C": "II", "D": "II", "E": "IV", "F": "V",
    "G": "VI", "H": "VII", "I": "IX", "J": "X", "K": "XI", "L": "XII",
    "M": "XIII", "N": "XIV", "O": "XV", "P": "XVI", "Q": "XVII",
    "R": "XVIII", "S": "XIX", "T": "XIX", "V": "XX", "W": "XX",
    "X": "XX", "Y": "XX", "Z": "XXI", "U": "XXII"
}

ICD9_DX_CHAPTERS = [
    (1, 139, "I", "Infectious and Parasitic Diseases"),
    (140, 239, "II", "Neoplasms"),
    (240, 279, "III", "Endocrine, Nutritional, and Metabolic Diseases"),
    (280, 289, "IV", "Diseases of the Blood"),
    (290, 319, "V", "Mental Disorders"),
    (320, 389, "VI", "Diseases of the Nervous System"),
    (390, 459, "VII", "Diseases of the Circulatory System"),
    (460, 519, "VIII", "Diseases of the Respiratory System"),
    (520, 579, "IX", "Diseases of the Digestive System"),
    (580, 629, "X", "Diseases of the Genitourinary System"),
    (630, 679, "XI", "Complications of Pregnancy"),
    (680, 709, "XII", "Diseases of the Skin"),
    (710, 739, "XIII", "Diseases of the Musculoskeletal System"),
    (740, 759, "XIV", "Congenital Anomalies"),
    (760, 779, "XV", "Perinatal Conditions"),
    (780, 799, "XVI", "Symptoms and Signs"),
    (800, 999, "XVII", "Injury and Poisoning"),
    (None, None, "XVIII", "Supplementary Factors (V-Codes)"),
    (None, None, "XIX", "Supplementary External Causes (E-Codes)"),
]

ICD9_PROC_CHAPTERS = [
    (0, 0, "00-00", "Procedures and Interventions, Not Elsewhere Classified"),
    (1, 5, "01-05", "Operations on the Nervous System"),
    (6, 7, "06-07", "Operations on the Endocrine System"),
    (8, 16, "08-16", "Operations on the Eye"),
    (17, 17, "17-17", "Other Miscellaneous Diagnostic and Therapeutic Procedures"),
    (18, 20, "18-20", "Operations on the Ear"),
    (21, 29, "21-29", "Operations on the Nose, Mouth, and Pharynx"),
    (30, 34, "30-34", "Operations on the Respiratory System"),
    (35, 39, "35-39", "Operations on the Cardiovascular System"),
    (40, 41, "40-41", "Operations on the Hemic and Lymphatic System"),
    (42, 54, "42-54", "Operations on the Digestive System"),
    (55, 59, "55-59", "Operations on the Urinary System"),
    (60, 64, "60-64", "Operations on the Male Genital Organs"),
    (65, 71, "65-71", "Operations on the Female Genital Organs"),
    (72, 75, "72-75", "Obstetrical Procedures"),
    (76, 84, "76-84", "Operations on the Musculoskeletal System"),
    (85, 86, "85-86", "Operations on the Integumentary System"),
    (87, 99, "87-99", "Miscellaneous Diagnostic and Therapeutic Procedures"),
]


def get_chapter_roman(clean_code: str) -> Optional[str]:
    if not clean_code or len(clean_code) < 1: return None
    letter = clean_code[0]
    if letter == "D": return "III" if (len(clean_code) > 1 and clean_code[1] in "56789") else "II"
    if letter == "H": return "VIII" if (len(clean_code) > 1 and clean_code[1] in "6789") else "VII"
    return ICD10_CHAPTER_LOOKUP.get(letter)


def get_icd9_dx_chapter(clean_code: str) -> Optional[str]:
    if not clean_code: return None
    if clean_code.startswith("V"): return "XVIII"
    if clean_code.startswith("E"): return "XIX"
    try:
        prefix = int(clean_code[:3])
        for start, end, roman, _ in ICD9_DX_CHAPTERS:
            if start is not None and start <= prefix <= end: return roman
    except ValueError:
        pass
    return None


def get_icd9_proc_chapter(clean_code: str) -> Optional[str]:
    if not clean_code: return None
    try:
        prefix = int(clean_code[:2])
        for start, end, key, _ in ICD9_PROC_CHAPTERS:
            if start <= prefix <= end: return key
    except ValueError:
        pass
    return None


def find_icd9_range_parent(target_code: str, all_keys: list, is_proc: bool = False) -> Optional[str]:
    """
    Finds the smallest ICD-9 numeric range containing the target code's prefix.
    If is_proc is True, consider 2-digit procedure prefixes as possible matches.
    Otherwise (diagnosis), only consider 3-digit category prefixes.
    """
    # --- Existing Exact Match Logic ---
    if len(target_code) >= 3:
        p3 = target_code[:3]
        if f"{p3}-{p3}" in all_keys:
            return f"{p3}-{p3}"

    # 1. Prepare Target Values
    match = re.match(r'^(\d+)', target_code)
    if not match:
        return None
    raw_digits = match.group(1)

    # Calculate potential comparison values
    target_vals = set()
    # For diagnoses prefer 3-digit category (if available)
    if len(raw_digits) >= 3:
        target_vals.add(int(raw_digits[:3]))
    # For procedures, also consider 2-digit blocks; for diagnoses we should NOT
    if is_proc and len(raw_digits) >= 2:
        target_vals.add(int(raw_digits[:2]))

    if not target_vals:
        return None

    # 2. Scan Blocks and find smallest container
    best_parent = None
    best_range_size = float('inf')

    for key in all_keys:
        if "-" not in key:
            continue
        parts = key.split("-")
        if len(parts) != 2 or not parts[0].isdigit():
            continue

        start, end = int(parts[0]), int(parts[1])

        # Check containment for ANY target value calculated
        for target_val in target_vals:
            if start <= target_val <= end:
                size = end - start
                # Prioritize the smallest valid range
                if size < best_range_size:
                    best_range_size = size
                    best_parent = key
                break  # Found a container for this block, move to next key

    return best_parent



def icd10_to_int(code: str) -> int:
    if not code or len(code) < 3: return -1
    letter = code[0].upper()
    if not letter.isalpha(): return -1
    try:
        return (ord(letter) - ord('A')) * 100 + int(code[1:3])
    except ValueError:
        return -1


def find_icd10_range_parent(target_range: str, all_keys: list) -> Optional[str]:
    if "-" not in target_range: return None
    parts = target_range.split("-")
    if len(parts) != 2: return None
    t_start = icd10_to_int(parts[0])
    t_end = icd10_to_int(parts[1])
    if t_start == -1 or t_end == -1: return None
    best_parent = None
    best_range_size = float('inf')
    for key in all_keys:
        if key == target_range: continue
        if "-" not in key: continue
        k_parts = key.split("-")
        if len(k_parts) != 2: continue
        p_start = icd10_to_int(k_parts[0])
        p_end = icd10_to_int(k_parts[1])
        if p_start == -1 or p_end == -1: continue

        if p_start <= t_start and p_end >= t_end:
            size = p_end - p_start
            if size < best_range_size:
                best_range_size = size
                best_parent = key
    return best_parent


def inject_icd9_chapters(graph, system_id, is_proc=False):
    source = ICD9_PROC_CHAPTERS if is_proc else ICD9_DX_CHAPTERS
    for _, _, code, desc in source:
        graph.nodes[code] = {
            "code": code, "description": desc, "parent_hint": None,
            "system_id": system_id, "is_selectable": False, "is_leaf": False,
            "db_id": None, "parent_db_id": None
        }


# ==========================================================
# 1. THE UNIFIED GRAPH
# ==========================================================
class UnifiedGraph:
    def __init__(self, family_name, root_code, root_desc):
        self.family_name = family_name
        self.root_code = root_code
        self.root_desc = root_desc
        self.nodes = {}
        self.base_system_id = None

    def add_system_data(self, codes: List[Dict], system_name: str, system_id: int):
        print(f"   🌱 Merging {len(codes)} codes from {system_name} into {self.family_name}...")
        if self.base_system_id is None: self.base_system_id = system_id

        count = 0
        for c in codes:
            key = c['clean_code']
            if key in self.nodes: continue
            desc = c['description']
            if "Base" not in system_name:
                if f"({system_name})" not in desc: desc = f"{desc} ({system_name})"

            self.nodes[key] = {
                "code": key, "description": desc, "parent_hint": c['parent_code'],
                "system_id": system_id, "is_selectable": c['is_selectable'],
                "is_leaf": c['is_leaf'], "db_id": None, "parent_db_id": None
            }
            count += 1
        print(f"      -> Added {count} new nodes.")

    def finalize_and_export(self, start_global_id: int):
        print(f"   🌳 Wiring the {self.family_name} tree...")
        root_sys_id = self.base_system_id if self.base_system_id else 1
        root_key = self.root_code

        # --- STEP 0: SYNTHESIZE MISSING PARENTS ---
        is_icd9 = "ICD-9" in self.family_name
        is_opcs = "OPCS" in self.family_name
        is_icd10 = "ICD-10" in self.family_name
        is_proc = "procedure" in self.family_name.lower()

        def run_synthesis_pass(target_len, parent_len):
            missing_map = {}
            current_keys = list(self.nodes.keys())
            for key in current_keys:
                if not key.isdigit(): continue
                if len(key) != target_len: continue
                needed = key[:parent_len]
                if needed not in self.nodes:
                    if needed not in missing_map: missing_map[needed] = []
                    missing_map[needed].append(self.nodes[key]['description'])
            created = 0
            for p_code, child_descs in missing_map.items():
                smart_desc = get_common_description(child_descs)
                self.nodes[p_code] = {
                    "code": p_code, "description": smart_desc,
                    "parent_hint": p_code[:3] if len(p_code) > 3 else None,
                    "system_id": root_sys_id, "is_selectable": False, "is_leaf": False,
                    "db_id": None, "parent_db_id": None
                }
                created += 1
            return created

        if (is_icd9 and not is_proc) or is_opcs:
            # Diagnosis / OPCS (5->4, 4->3)
            c1 = run_synthesis_pass(5, 4)
            c2 = run_synthesis_pass(4, 3)
            if c1 + c2 > 0: print(f"      👻 Synthesized {c1 + c2} missing Dx parents.")

        if is_icd9 and is_proc:
            # Procedures (4->3, 3->2)
            # We need strict 2-digit roots for Proc
            c1 = run_synthesis_pass(4, 3)
            c2 = run_synthesis_pass(3, 2)
            if c1 + c2 > 0: print(f"      👻 Synthesized {c1 + c2} missing Proc parents.")

        # 1. Root
        self.nodes[root_key] = {
            "code": root_key, "description": self.root_desc, "parent_hint": None,
            "system_id": root_sys_id, "is_selectable": False, "is_leaf": False,
            "db_id": start_global_id, "parent_db_id": None,
            "materialized_path": f"/{start_global_id}/"
        }
        curr_id = start_global_id + 1

        keys_to_sort = [k for k in self.nodes if k != root_key]
        sorted_keys = sorted(keys_to_sort, key=custom_sort_key)
        for key in sorted_keys:
            self.nodes[key]["db_id"] = curr_id
            curr_id += 1

        # 3. Resolve Parents
        all_keys = list(self.nodes.keys())

        for key in sorted_keys:
            node = self.nodes[key]
            hint = normalize_code(node['parent_hint'])
            p_node = None

            # DEBUG: Track what's happening with 1901 in OPCS-3
            debug_this = (is_opcs and key in ['1901', '190'])
            if debug_this:
                print(f"\n🐛 DEBUGGING {self.family_name}: {key}")
                print(f"   Hint: '{hint}'")
                print(f"   Key length: {len(key)}, is_digit: {key.isdigit()}")

            # --- 1. Specific Parent Heuristic (IMPROVED) ---

            # A. Try String Slicing FIRST (for numeric codes that should have numeric parents)
            if (is_icd9 and is_proc) or is_opcs:
                if len(key) > 2 and key.isdigit():
                    imp_hint = key[:-1]
                    if debug_this:
                        print(f"   Trying slice FIRST: '{key}' -> '{imp_hint}'")
                        print(f"   imp_hint in nodes: {imp_hint in self.nodes}")
                    if imp_hint in self.nodes:
                        p_node = self.nodes[imp_hint]
                        if debug_this: print(f"   ✓ Found via slice: {p_node['code']}")

            # B. Try Explicit Hint (only if slicing didn't work)
            if not p_node and hint and hint in self.nodes:
                p_node = self.nodes[hint]
                if debug_this: print(f"   ✓ Found via hint: {p_node['code']}")

            # --- 2. Range Heuristics ---
            # Only fallback to range if no specific parent found

            # C1. ICD-10 Super-Block Override
            if is_icd10 and "-" in key:
                super_block = find_icd10_range_parent(key, all_keys)
                if super_block: p_node = self.nodes[super_block]

            # C2. ICD-9 Range
            if not p_node and is_icd9:
                # If 3-digits (042), looks for 042-044
                # If 2-digits Proc (60), looks for 60-64
                if key.isdigit():
                    r_key = find_icd9_range_parent(key, all_keys, is_proc=is_proc)
                    if r_key and r_key != key: p_node = self.nodes[r_key]

            # D. Chapter Heuristic
            if not p_node and key not in ROMAN_MAP:
                ch = None
                if is_icd10:
                    ch = get_chapter_roman(key)
                elif is_icd9 and not is_proc:
                    ch = get_icd9_dx_chapter(key)
                elif is_icd9 and is_proc:
                    ch = get_icd9_proc_chapter(key)

                if ch and ch in self.nodes and ch != key:
                    p_node = self.nodes[ch]

            # E. Fallback to Root
            if not p_node: p_node = self.nodes[root_key]
            if p_node['db_id'] == node['db_id']: p_node = self.nodes[root_key]

            p_node['is_leaf'] = False
            node['parent_db_id'] = p_node['db_id']

        # 4. Selectability
        parent_db_ids = set(n['parent_db_id'] for n in self.nodes.values() if n['parent_db_id'])
        for key in sorted_keys:
            node = self.nodes[key]
            is_active_parent = (node['db_id'] in parent_db_ids)
            node['is_leaf'] = not is_active_parent

            if is_active_parent:
                is_structure = (key in ROMAN_MAP) or ("-" in key) or ("CPT" in self.family_name)
                if is_structure:
                    node['is_selectable'] = False
                else:
                    node['is_selectable'] = True
                    if not node['code'].endswith("*"):
                        node['code'] += "*"

        # 5. Paths (Cycle-Safe)
        id_to_node = {n['db_id']: n for n in self.nodes.values()}
        path_cache = {}
        processing = set()

        def resolve_path(current_db_id):
            if current_db_id in path_cache: return path_cache[current_db_id]
            if current_db_id in processing:
                root_n = self.nodes[root_key]
                id_to_node[current_db_id]['parent_db_id'] = root_n['db_id']
                return f"/{root_n['db_id']}/{current_db_id}/"

            processing.add(current_db_id)
            node = id_to_node[current_db_id]
            pid = node['parent_db_id']

            if pid and pid in id_to_node and pid != current_db_id:
                parent_path = resolve_path(pid)
            else:
                root_n = self.nodes[root_key]
                parent_path = "/" if node['code'] == self.root_code else f"/{root_n['db_id']}/"

            my_path = f"{parent_path}{current_db_id}/"
            processing.remove(current_db_id)
            path_cache[current_db_id] = my_path
            return my_path

        path_cache[self.nodes[root_key]['db_id']] = f"/{start_global_id}/"
        for key in sorted_keys:
            node = self.nodes[key]
            node['materialized_path'] = resolve_path(node['db_id'])

        out = [self._fmt(self.nodes[root_key])]
        for key in sorted_keys: out.append(self._fmt(self.nodes[key]))
        return out, curr_id

    def _fmt(self, n):
        return {
            "id": n['db_id'], "system_id": n['system_id'], "code": n['code'],
            "description": n['description'], "parent_id": n['parent_db_id'],
            "materialized_path": n['materialized_path'],
            "is_leaf": n['is_leaf'], "is_selectable": n['is_selectable']
        }


# ==========================================================
# 2. PARSERS
# ==========================================================

def parse_ukbb_tree(file_path):
    if not os.path.exists(file_path): return []
    id_map = {}
    rows = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter='\t')
        for row in reader:
            clean = normalize_code(row['coding'])
            id_map[row['node_id']] = clean
            rows.append(row)
    codes = []
    for row in rows:
        clean = normalize_code(row['coding'])
        clean_desc = strip_code_from_desc(row['coding'], row['meaning'])
        p_code = id_map.get(row['parent_id'])
        codes.append({
            "clean_code": clean, "description": clean_desc, "parent_code": p_code,
            "is_leaf": True, "is_selectable": (row['selectable'] == 'Y')
        })
    return codes


def parse_icd10_who(file_path):
    if not os.path.exists(file_path): return []
    tree = ET.parse(file_path)
    codes = [];
    stack = []
    for cls in tree.getroot().findall(".//Class"):
        raw = cls.attrib.get("code");
        kind = cls.attrib.get("kind")
        clean = normalize_code(raw)
        desc = cls.find(".//Rubric/Label").text or ""
        parent = None
        if kind == "chapter":
            stack = []
        elif kind == "block":
            stack.append(clean)
        elif kind == "category":
            parent = normalize_code(raw.split(".")[0]) if "." in raw else (stack[-1] if stack else None)
        codes.append({
            "clean_code": clean, "description": strip_code_from_desc(raw, desc), "parent_code": parent,
            "is_leaf": True, "is_selectable": (kind == "category")
        })
    return codes


def parse_icd10_cm(file_path):
    if not os.path.exists(file_path): return []
    tree = ET.parse(file_path)
    codes = []

    def process(el, p):
        raw = el.find('name').text;
        clean = normalize_code(raw)
        codes.append({
            "clean_code": clean, "description": el.find('desc').text,
            "parent_code": p, "is_leaf": True, "is_selectable": True
        })
        for ch in el.findall('diag'): process(ch, clean)

    for sec in tree.getroot().findall('.//section'):
        for diag in sec.findall('diag'): process(diag, None)
    return codes


def parse_icd9_cms(file_path, is_proc=False):
    if not os.path.exists(file_path): return []
    codes = []
    with open(file_path, 'r', encoding='latin-1') as f:
        for line in f:
            parts = line.strip().split(maxsplit=1)
            if len(parts) < 2: continue
            raw = parts[0];
            desc = parts[1]
            clean = normalize_code(raw)
            clean_desc = strip_code_from_desc(raw, desc)

            threshold = 2 if is_proc else 3
            parent = clean[:-1] if len(clean) > threshold else None

            codes.append({
                "clean_code": clean, "description": clean_desc, "parent_code": parent,
                "is_leaf": True, "is_selectable": True
            })
    return codes


def parse_cpt_excel(file_path):
    """
    Parses CPT codes into a flat 2-level hierarchy.
    Level 1: Category (Code = Category ID, Desc = Operative Procedure)
    Level 2: CPT Code (Code = CPT Code, Desc = Procedure Desc)
    """
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return []

    print(f"   ⚙️ Parsing CPT Excel: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"   ❌ Error opening Excel: {e}")
        return []

    codes_list = []
    seen_categories = set()

    # Helper to handle header scanning
    def get_header_map(worksheet):
        rows = list(worksheet.iter_rows(min_row=1, max_row=20))
        for r_idx, row in enumerate(rows):
            values = [str(c.value).strip().lower() for c in row if c.value]
            if not values: continue

            h_map = {}
            for i, cell in enumerate(row):
                if cell.value:
                    val = str(cell.value).strip().lower().replace("\n", " ").replace("  ", " ")
                    h_map[val] = i

            if "procedure code category" in h_map:
                return h_map, r_idx + 1
        return None, None

    # -------------------------------------------------------
    # PASS 1: Build Map from "Index" Tab
    # Map: Category ID (Code) -> Operative Procedure (Description)
    # -------------------------------------------------------
    cat_desc_map = {}

    if "Index" in wb.sheetnames:
        ws_index = wb["Index"]
        idx_headers, idx_start = get_header_map(ws_index)

        if idx_headers:
            c_cat = idx_headers.get("procedure code category")
            c_op = idx_headers.get("operative procedure")

            for row in ws_index.iter_rows(min_row=idx_start + 1, values_only=True):
                raw_cat = row[c_cat]
                raw_desc = row[c_op]
                if raw_cat and raw_desc:
                    # Key = "10004-10021" (Code)
                    # Value = "Surgery" (Description)
                    cat_desc_map[str(raw_cat).strip()] = str(raw_desc).strip()

    # -------------------------------------------------------
    # PASS 2: Codes Tab (Build the Tree)
    # -------------------------------------------------------
    target_tab = "ALL 2024 CPT Codes"
    if target_tab not in wb.sheetnames:
        found = [s for s in wb.sheetnames if "CPT Codes" in s]
        if found:
            target_tab = found[0]
        else:
            return []

    ws_codes = wb[target_tab]
    code_headers, code_start = get_header_map(ws_codes)

    if not code_headers:
        print("   ❌ CPT Headers not found.")
        return []

    idx_code = code_headers.get("cpt codes")
    idx_desc = code_headers.get("procedure code descriptions")
    idx_cat = code_headers.get("procedure code category")

    for row in ws_codes.iter_rows(min_row=code_start + 1, values_only=True):
        raw_code = row[idx_code]
        if not raw_code: continue

        # 1. Clean Child Code
        clean_code = str(raw_code).strip()
        if clean_code.isdigit() and len(clean_code) < 5:
            clean_code = clean_code.zfill(5)

        # 2. Identify Parent (Category)
        raw_cat = row[idx_cat]
        clean_cat = str(raw_cat).strip() if raw_cat else "Uncategorized"
        clean_cat_norm = normalize_code(clean_cat)  # This will be the Parent Code

        # 3. Create Parent Node (If new)
        if clean_cat_norm not in seen_categories:
            # Lookup the Description we mapped in Pass 1
            parent_desc = cat_desc_map.get(clean_cat, clean_cat)

            codes_list.append({
                "clean_code": clean_cat_norm,  # Code = "10004-10021"
                "description": parent_desc,  # Desc = "Surgery"
                "parent_code": None,  # Attached directly to Root
                "is_leaf": False,
                "is_selectable": False  # Categories are folders
            })
            seen_categories.add(clean_cat_norm)

        # 4. Add Child Node
        desc = row[idx_desc] if idx_desc is not None else ""

        codes_list.append({
            "clean_code": clean_code,
            "description": str(desc).strip(),
            "parent_code": clean_cat_norm,
            "is_leaf": True,
            "is_selectable": True
        })

    print(f"      -> Parsed {len(codes_list)} CPT entities.")
    return codes_list


# ==========================================================
# 3. DEBUG
# ==========================================================
def print_debug_tree(rows, start_code, depth=3, system_filter=None):
    children_map = {}
    for r in rows:
        pid = r['parent_id']
        if pid: children_map.setdefault(pid, []).append(r)

    search = normalize_code(start_code)
    matches = [r for r in rows if r['code'] == search or r['code'] == f"{search}*"]
    if system_filter: matches = [r for r in matches if r['system_id'] == system_filter]

    if not matches: print(f"❌ '{search}' not found."); return

    for root in matches:
        print(f"\n🔎 ROOT: {root['code']} (ID: {root['id']}, Sys: {root['system_id']})")

        def _print(node, d, pre):
            if d > depth: return
            kids = children_map.get(node['id'], [])
            kids.sort(key=lambda x: x['code'])
            if not kids and d == 1: print(f"   ⚠️ Empty.")
            for i, k in enumerate(kids):
                last = (i == len(kids) - 1)
                icon = "🟢" if k['is_selectable'] else "📁"
                print(f"{pre}{'└── ' if last else '├── '}{icon} [{k['code']}] {k['description'][:60]}")
                _print(k, d + 1, pre + ("    " if last else "│   "))

        _print(root, 1, "")


# ==========================================================
# 4. RUNNER
# ==========================================================
def run_export():
    base_dir = os.path.expanduser("~/Downloads")
    files = {
        "who": f"{base_dir}/icd102019en.xml",
        "cm": f"{base_dir}/icd10cm-table and index-2026/icd10cm-tabular-2026.xml",
        "ukbb10": f"{base_dir}/coding19.tsv",
        "icd9": f"{base_dir}/ICD-9-CM-v32-master-descriptions/CMS32_DESC_LONG_DX.txt",
        "ukbb9": f"{base_dir}/coding87.tsv",
        "opcs3": f"{base_dir}/coding259.tsv",
        "opcs4": f"{base_dir}/coding240.tsv",
        "icd9proc": f"{base_dir}/ICD-9-CM-v32-master-descriptions/CMS32_DESC_LONG_SG.txt",
        "cpt": f"{base_dir}/cpt-pcm-nhsn.xlsx"
    }

    next_sys = 1;
    next_gid = 1;
    systems = [];
    codes = []

    # --- 1. ICD-10 ---
    g10 = UnifiedGraph("ICD-10 unified", "ICD-10", "Unified ICD-10 codes (WHO, CMS, UK Biobank)")
    s1 = {"id": next_sys, "name": "ICD-10-UKBB", "description": "ICD-10 codes provided through the UK Biobank data dictionary", "version": "Accessed November 2025", "url": "https://biobank.ndph.ox.ac.uk/ukb/coding.cgi?id=19"};
    systems.append(s1)
    g10.add_system_data(parse_ukbb_tree(files["ukbb10"]), "UKBB (Base)", next_sys);
    next_sys += 1
    s2 = {"id": next_sys, "name": "ICD-10-WHO", "description": "ICD-10 2019 version (including COVID-19 updates)", "version": "2019", "url": "https://icdcdn.who.int/icd10/claml/icd102019en.xml.zip"};
    systems.append(s2)
    g10.add_system_data(parse_icd10_who(files["who"]), "ICD-10-WHO", next_sys);
    next_sys += 1
    s3 = {"id": next_sys, "name": "ICD-10-CM", "description": "ICD-10 Clinical Modification", "version": "2026", "url": "https://ftp.cdc.gov/pub/health_statistics/nchs/publications/ICD10CM/2026/icd10cm-table%20and%20index-2026.zip"};
    systems.append(s3)
    g10.add_system_data(parse_icd10_cm(files["cm"]), "ICD-10-CM", next_sys);
    next_sys += 1
    r10, next_gid = g10.finalize_and_export(next_gid)
    codes.extend(r10)

    # --- 2. ICD-9 ---
    g9 = UnifiedGraph("ICD-9 unified", "ICD-9", "Unified ICD-9 codes (CMS, UK Biobank)")
    s4 = {"id": next_sys, "name": "ICD-9-UKBB", "description": "ICD-9 codes provided through the UK Biobank data dictionary", "version": "Accessed November 2025", "url": "https://biobank.ctsu.ox.ac.uk/ukb/coding.cgi?id=87"};
    systems.append(s4)
    inject_icd9_chapters(g9, next_sys, is_proc=False)
    g9.add_system_data(parse_ukbb_tree(files["ukbb9"]), "UKBB (Base)", next_sys);
    next_sys += 1
    s5 = {"id": next_sys, "name": "ICD-9-CM", "description": "ICD-9 Clinical Modification", "version": "2014 (v32)", "url": "https://www.cms.gov/medicare/coding/icd9providerdiagnosticcodes/downloads/icd-9-cm-v32-master-descriptions.zip"};
    systems.append(s5)
    g9.add_system_data(parse_icd9_cms(files["icd9"]), "ICD-9-CM", next_sys);
    next_sys += 1
    r9, next_gid = g9.finalize_and_export(next_gid)
    codes.extend(r9)

    # --- 3. OPCS-4 ---
    g_opcs4 = UnifiedGraph("OPCS-4 Unified", "OPCS-4", "OPCS-4 Procedure Codes")
    s6 = {"id": next_sys, "name": "OPCS-4-UKBB", "description": "OPSC4 codes provided through the UK Biobank data dictionary", "version": "Accessed November 2025", "url": "https://biobank.ndph.ox.ac.uk/ukb/coding.cgi?id=240"};
    systems.append(s6)
    g_opcs4.add_system_data(parse_ukbb_tree(files["opcs4"]), "UKBB (Base)", next_sys);
    next_sys += 1
    r_opcs4, next_gid = g_opcs4.finalize_and_export(next_gid)
    codes.extend(r_opcs4)

    # --- 4. OPCS-3 ---
    g_opcs3 = UnifiedGraph("OPCS-3 Unified", "OPCS-3", "OPCS-3 Procedure Codes")
    s7 = {"id": next_sys, "name": "OPCS-3-UKBB", "description": "OPSC4 codes provided through the UK Biobank data dictionary", "version": "Accessed November 2025", "url": "https://biobank.ndph.ox.ac.uk/ukb/coding.cgi?id=259"};
    systems.append(s7)
    g_opcs3.add_system_data(parse_ukbb_tree(files["opcs3"]), "UKBB (Base)", next_sys);
    next_sys += 1
    r_opcs3, next_gid = g_opcs3.finalize_and_export(next_gid)
    codes.extend(r_opcs3)

    # --- 5. ICD-9 Proc ---
    g9p = UnifiedGraph("ICD-9 procedures unified", "ICD-9-Proc", "ICD-9 Procedures Codes")
    s8 = {"id": next_sys, "name": "ICD-9-CM-Proc", "description": "ICD-9 Procedure Codes", "version": "2014 (v32)", "url": "https://www.cms.gov/medicare/coding/icd9providerdiagnosticcodes/downloads/icd-9-cm-v32-master-descriptions.zip"};
    systems.append(s8)
    inject_icd9_chapters(g9p, next_sys, is_proc=True)
    g9p.add_system_data(parse_icd9_cms(files["icd9proc"], is_proc=True), "ICD-9-Proc", next_sys);
    next_sys += 1
    r9p, next_gid = g9p.finalize_and_export(next_gid);
    codes.extend(r9p)

    # --- 6. CPT ---
    g_cpt = UnifiedGraph("CPT-4 Unified", "CPT-4", "Current Procedural Terminology Codes")
    s9 = {"id": next_sys, "name": "CPT-4", "description": "2024 NHSN CPT Operative Procedure Code Mappings (updated 1/2024)", "version": "2024", "url": "https://www.cdc.gov/nhsn/pdfs/validation/2024/opc-cpt-pcm-nhsn.xlsx"};
    systems.append(s9)
    g_cpt.add_system_data(parse_cpt_excel(files["cpt"]), "CPT", next_sys);
    next_sys += 1
    r_cpt, next_gid = g_cpt.finalize_and_export(next_gid);
    codes.extend(r_cpt)

    print("\n💾 Saving...")
    codes.sort(key=lambda x: x['id'])
    with open(f"{base_dir}/code_systems.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["id", "name", "description", "version", "url"]);
        w.writeheader();
        w.writerows(systems)
    with open(f"{base_dir}/codes.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["id", "system_id", "code", "description", "parent_id", "materialized_path",
                                          "is_leaf", "is_selectable"]);
        w.writeheader();
        w.writerows(codes)

    print(f"🎉 Done! Created {len(codes)} unified nodes.")

    # DEBUG
    print("\n🔍 CHECKING 59* SERIES (Proc):")
    print_debug_tree(codes, start_code="59", depth=2)


if __name__ == "__main__":
    run_export()